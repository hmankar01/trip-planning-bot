from openpyxl import load_workbook

wb_flight = load_workbook("data/flights.xlsx")
wb_hotel = load_workbook("data/hotels.xlsx")
wb_place = load_workbook("data/places.xlsx")


# Flight details


def give_flight_options():
    sheet = wb_flight.active
    cities = set()
    city_dict = {}

    def print_options(loc):
        print(f'Choose your {loc} city:')
        for cell in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1, values_only=True):
            cities.add(cell[0])

        for count, city in enumerate(cities, start=1):
            city_dict.update({count: city})

        for option in city_dict:
            print(f'{option}){city_dict[option]}')
        try:
            option_number = int(input('> '))
            return city_dict[option_number]
        except KeyError:
            print('Please choose from amongst the given numbers.')
            print_options(loc)

    user_location = print_options('starting')
    user_destination = print_options('destination')

    if user_location == user_destination:
        print('The starting  city cannot be the same as the destination city, please try again.')
        main()
    return user_location, user_destination


def get_flight_price(curr_budget):
    sheet = wb_flight.active
    user_location, user_destination = give_flight_options()
    row = 1

    for cell in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=3, values_only=True):
        row += 1
        if cell[0] == user_location and cell[1] == user_destination:
            price = cell[2]
            print(f'Flight from {user_location} to {user_destination} costs {price} Rs.')
            if curr_budget < price:
                budget_warning('flights')
            else:
                return user_destination, price

# Hotel details


def get_hotel_price(destination_city, curr_budget):
    sheet = wb_hotel.active
    hotel_dict = {}
    hotels = []
    row = 1
    for cell in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=3, values_only=True):
        row += 1
        if cell[0] == destination_city and int(cell[2]) < curr_budget:
            hotels.append([cell[1], cell[2]])
    if len(hotels) == 0:
        budget_warning('hotels')

    def print_options():
        print(f'Choose your hotel:')
        for count, hotel in enumerate(hotels, start=1):
            hotel_dict.update({count: hotel})

        for i, option in enumerate(hotel_dict, start=1):
            print(f'{i}){(hotel_dict[i][0])}------{(hotel_dict[i][1])} Rs')
            i += 1
        try:
            option_number = int(input('> '))
            return hotel_dict[option_number][0], hotel_dict[option_number][1]
        except KeyError:
            print('Please choose from amongst the given numbers.')
            print_options()
    user_hotel, price = print_options()
    print(f'Your stay at {user_hotel} will cost you {price} Rs.')
    return price


# Place details


def get_place_price(destination_city, curr_budget):
    sheet = wb_place.active
    place_dict = {}
    places = []
    row = 1
    for cell in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=3, values_only=True):
        row += 1
        if cell[0] == destination_city and int(cell[2]) < curr_budget:
            places.append([cell[1], cell[2]])
    if len(places) == 0:
        budget_warning('places')

    def print_options():
        print(f'Choose the places to visit:')
        for count, place in enumerate(places, start=1):
            place_dict.update({count: place})

        for i, option in enumerate(place_dict, start=1):
            print(f'{i}){(place_dict[i][0])}------{(place_dict[i][1])} Rs')
            i += 1
        try:
            option_number = int(input('> '))
            return place_dict[option_number][0], place_dict[option_number][1]
        except KeyError:
            print('Please choose from amongst the given numbers.')
            print_options()
    user_place, price = print_options()
    print(f'Your trip to {user_place} will cost you {price} Rs.')
    return price


# Extra functions

def budget_warning(problem):
    print(f"You don't have any {problem} available in your budget")
    print("Please revaluate your budget.\n")
    main()
    exit()


def calc_budget(budget, cost):
    budget -= cost
    print(f'Your remaining budget is {budget} Rs.\n')
    return budget


def main():
    budget = int(input("What is your budget?: "))
    user_destination, flight_price = get_flight_price(budget)
    budget = calc_budget(budget, flight_price)
    hotel_price = get_hotel_price(user_destination, budget)
    budget = calc_budget(budget, hotel_price)
    place_price = get_place_price(user_destination, budget)
    calc_budget(budget, place_price)


if __name__ == "__main__":
    main()
